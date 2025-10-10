import os
import json
import subprocess
import shutil
import time
from tqdm import tqdm
from bs4 import BeautifulSoup
import pandas as pd
import json
import traceback
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from parse_html import parse_html
from parse_json import parse_tweet_json
from sec_downloads import fetch_html_after_delay,extract_redirect_on_302
from pathlib import Path
from urllib.parse import urlparse
from internet import has_internet 

def secondary_download(url,timestamp,project_dir,output_dir):
    try:
        link = f"https://web.archive.org/web/{timestamp}/{url}"
        safe_name = url.replace("https://", "").replace("http://", "").replace("/", "_")
        ext = ".html"
        file_name = f"{timestamp}_{safe_name}{ext}"
        final_file = os.path.join(output_dir, file_name)

        content = fetch_html_after_delay(link)
        redirect_link = extract_redirect_on_302(content)
        if redirect_link:
            content = fetch_html_after_delay(redirect_link)
            retweet = True
        else:
            retweet = False
        with open(final_file,"w" ,encoding="utf-8") as file:
            file.write(content)

        if "<!DOCTYPE html>" in content:
            soup = BeautifulSoup(content, "html.parser")
            pre_tag = soup.select_one("div#jsonview pre")

            if pre_tag:
                text_content = pre_tag.get_text(strip=True)
                # print(text_content)  # 👉 {"hello": "world"}
                tweet_json = json.loads(text_content)
                tweet = parse_tweet_json(tweet_json, timestamp)
                print(tweet)
                update_xlsx(project_dir=project_dir,tweet=tweet)
                
            else:
                tweet = parse_html(soup, link, True ,retweet)
                print(tweet)
                update_xlsx(project_dir=project_dir,tweet=tweet)

                
        else:
            tweet_json = json.loads(content)
            tweet = parse_tweet_json(tweet_json, timestamp)
            print(tweet)
            update_xlsx(project_dir=project_dir,tweet=tweet)
        return True
    except Exception as e:
        print(f"Error processing row: {e}")
        # traceback.print_exc()
        
        error_message = traceback.format_exc()
        print(error_message)
        try:
            with open(os.path.join(project_dir,f"{project_dir}_error_tweets.txt"), "r", encoding="utf-8") as f:
                error_tweets = f.read().splitlines()
        except FileNotFoundError:
            error_tweets = []

        # Use a set for faster lookups
        error_urls = set(error_tweets)

        # Append only if not already in log
        if url not in error_urls:
            error_urls.add(url)
            error_tweets.append(url)
            error_tweets.append(error_message)

            with open(os.path.join(project_dir,f"{project_dir}_error_tweets.txt"), "w", encoding="utf-8") as f:
                f.write("\n".join(error_tweets))
        return False


def delete_folder(folder_path):
    # Check if the folder exists before attempting to delete it
    if os.path.isdir(folder_path):
        try:
            shutil.rmtree(folder_path)
            print(f"Folder '{folder_path}' and its contents deleted successfully.")
        except OSError as e:
            print(f"Error: {folder_path} : {e.strerror}")
    else:
        print(f"Folder '{folder_path}' does not exist.")


def update_errors_xlsx(project_dir,tweet,file_number):
    csv_file_path = os.path.join(project_dir,f'{project_dir}_tweets_fix_{file_number}.csv')
    excel_file_path = os.path.join(project_dir,f'{project_dir}_tweets_fix_{file_number}.xlsx')
    """
    Updates the CSV with a new tweet and regenerates the formatted Excel file.
    
    tweet: dict
        Example:
        {
            "tweet_text": "Some tweet text",
            "date": "2020-09-03 08:04:00",
            "image": True,
            "quote": False,
            "reply": True,
            "mentions": "@someone",
            "username": "user123",
            "link": "https://twitter.com/user123/status/..."
        }
    """

    # Ensure CSV exists, otherwise create with headers
    file_exists = os.path.isfile(csv_file_path)
    df_new = pd.DataFrame([tweet])

    if file_exists:
        df = pd.read_csv(csv_file_path)
        df = pd.concat([df, df_new], ignore_index=True)

        # 🔑 Remove duplicates (keep first occurrence)
        df = df.drop_duplicates(subset=["link"], keep="first").reset_index(drop=True)
    else:
        df = df_new

    # Reorder columns
    desired_order = ["tweet_text", "date", "image", "quote", "reply", "mentions","retweet","author", "username", "link"]
    df = df[desired_order]

    # Save updated CSV
    df.to_csv(csv_file_path, index=False)

    # Save to Excel first
    df.to_excel(excel_file_path, index=False)

    # Load workbook
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Define fills (colors)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

    # Format columns
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        col_values = df[col_name].astype(str)
        max_len = max(col_values.map(len).max(), len(col_name))

        # Special case for tweet_text
        if col_name == "tweet_text":
            desired_width = max_len // 3 + 5
            wrap = True
        else:
            desired_width = min(max_len + 5, 50)  # cap width
            wrap = False

        ws.column_dimensions[col_letter].width = desired_width

        # Apply cell formatting
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=wrap, vertical="bottom")

                # Color logic
                if col_name == "image" and str(cell.value).upper() == "TRUE":
                    cell.fill = green_fill
                elif col_name == "quote" and str(cell.value).upper() == "TRUE":
                    cell.fill = yellow_fill
                elif col_name == "reply" and str(cell.value).upper() == "TRUE":
                    cell.fill = blue_fill

    # Save workbook
    wb.save(excel_file_path)

    print(f"Tweet added and '{excel_file_path}' updated successfully with formatting.")




def update_xlsx(project_dir,tweet):
    csv_file_path = os.path.join(project_dir,f'{project_dir}_tweets.csv')
    excel_file_path = os.path.join(project_dir,f'{project_dir}_tweets.xlsx')
    """
    Updates the CSV with a new tweet and regenerates the formatted Excel file.
    
    tweet: dict
        Example:
        {
            "tweet_text": "Some tweet text",
            "date": "2020-09-03 08:04:00",
            "image": True,
            "quote": False,
            "reply": True,
            "mentions": "@someone",
            "username": "user123",
            "link": "https://twitter.com/user123/status/..."
        }
    """

    # Ensure CSV exists, otherwise create with headers
    file_exists = os.path.isfile(csv_file_path)
    df_new = pd.DataFrame([tweet])

    if file_exists:
        df = pd.read_csv(csv_file_path)
        df = pd.concat([df, df_new], ignore_index=True)

        # 🔑 Remove duplicates (keep first occurrence)
        df = df.drop_duplicates(subset=["link"], keep="first").reset_index(drop=True)
    else:
        df = df_new

    # Reorder columns
    desired_order = ["tweet_text", "date", "image", "quote", "reply", "mentions","retweet","author", "username", "link"]
    df = df[desired_order]

    # Save updated CSV
    df.to_csv(csv_file_path, index=False)

    # Save to Excel first
    df.to_excel(excel_file_path, index=False)

    # Load workbook
    wb = load_workbook(excel_file_path)
    ws = wb.active

    # Define fills (colors)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

    # Format columns
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        col_values = df[col_name].astype(str)
        max_len = max(col_values.map(len).max(), len(col_name))

        # Special case for tweet_text
        if col_name == "tweet_text":
            desired_width = max_len // 3 + 5
            wrap = True
        else:
            desired_width = min(max_len + 5, 50)  # cap width
            wrap = False

        ws.column_dimensions[col_letter].width = desired_width

        # Apply cell formatting
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=wrap, vertical="bottom")

                # Color logic
                if col_name == "image" and str(cell.value).upper() == "TRUE":
                    cell.fill = green_fill
                elif col_name == "quote" and str(cell.value).upper() == "TRUE":
                    cell.fill = yellow_fill
                elif col_name == "reply" and str(cell.value).upper() == "TRUE":
                    cell.fill = blue_fill

    # Save workbook
    wb.save(excel_file_path)

    print(f"Tweet added and '{excel_file_path}' updated successfully with formatting.")

def download_file(url, timestamp, project_dir, output_dir, user_name):
    # Temporary folder for downloader
    first_download = True
    tmp_dir = os.path.abspath(os.path.join(output_dir, "tmp_dl"))

    if os.path.exists(tmp_dir):
        shutil.rmtree(tmp_dir)  # clean old runs
    os.makedirs(tmp_dir, exist_ok=True)

    # Run wayback_machine_downloader
    cmd = [
        "ruby",
        "wayback_machine_downloader",
        url,
        "-e",
        "-d",
        f"../../{user_name}/{project_dir}_archive/tmp_dl",
        "-f",
        timestamp
    ]
    print(f"Running: {cmd}")
    cwd = "wayback-machine-downloader/bin"

    try:
        result = subprocess.run(
            cmd,
            cwd=cwd,
            shell=True,             # needed since cmd is a string
            check=True,
            capture_output=True,    # capture stdout + stderr
            text=True               # decode to str instead of bytes
        )

        # Check if the output contains "No files to download"
        output = result.stdout + result.stderr
        if "No files to download." in output:
            print("⚠️ No files to download. The site may not be in Wayback Machine.")
            print("Trying secondary download ...")
            download_successful = secondary_download(
                url=url,
                timestamp=timestamp,
                project_dir=project_dir,
                output_dir=output_dir
            )
            shutil.rmtree(tmp_dir, ignore_errors=True)
            if download_successful:
                print("Secondary download completed successfully.")
            else:
                print("Secondary download failed")
            return None
        print("✅ Download completed successfully.")
    except subprocess.CalledProcessError as e:
        print(f"❌ Downloader failed for {url}: {e}")
        if e.stdout:
            print("STDOUT:", e.stdout)
        if e.stderr:
            print("STDERR:", e.stderr)
        shutil.rmtree(tmp_dir, ignore_errors=True)
        return None

    # Find the downloaded file (index.html or index.json)
    path_parts = urlparse(url).path.split("/")

    # Username is the part after "twitter.com/"
    user_name = path_parts[1] if len(path_parts) > 1 else "unknown"

    open_temp_dir = os.path.join(tmp_dir, user_name, "status")
    files = os.listdir(open_temp_dir)
    file = files[0]
    index_file = os.listdir(f"{open_temp_dir}/{file}")[0]
    index_path = os.path.join(open_temp_dir, file, index_file)

    return index_path,open_temp_dir




def download_with_wmd(url, timestamp,project_dir,user_name, content_type="text/html", output_dir="archive",update_errors=False,file_number="1"):
    os.makedirs(output_dir, exist_ok=True)

    safe_name = url.replace("https://", "").replace("http://", "").replace("/", "_")
    ext = ".html"
    file_name = f"{timestamp}_{safe_name.split("?")[0]}{ext}"
    final_file = os.path.join(output_dir, file_name)

    if os.path.exists(os.path.join(project_dir,f"{user_name}_tweets.csv")):
        try:
            df = pd.read_csv(os.path.join(project_dir,f"{user_name}_tweets.csv"))
            if "link" in df.columns and url in df["link"].values:
                print(f"Skipping (already in tweets.csv): {url}")
                return False
        except Exception as e:
            print(f"Warning: Could not check tweets.csv properly: {e}")
    
    # Skip if already saved
    if os.path.exists(final_file):
        print(f"Skipping (already saved): {final_file}")
        index_path = final_file
        if update_errors == False:
            return False
        first_download = False
    else:
        first_download = True
        index_path,open_temp_dir = download_file(url=url,timestamp=timestamp,project_dir=project_dir,output_dir=output_dir,user_name=user_name)
    
    with open(index_path, encoding="utf-8") as f:
        data = f.read()

    # Save permanent copy
    if first_download:
        with open(final_file, "w", encoding="utf-8") as f:
            f.write(data)
            # print(data)

        # Clean temp
        try:
            delete_folder(open_temp_dir)
            print(f"Folder '{open_temp_dir}' deleted successfully.")
        except OSError as e:
            print(f"Error deleting folder '{open_temp_dir}': {e}")

    try:
        # Parse according to type
        if content_type == "application/json":
            # tweet = parse_tweet_json(json.loads(data), timestamp=timestamp)
            if "<!DOCTYPE html>" in data:
                link = f"https://web.archive.org/web/{timestamp}/{url}"
                soup = BeautifulSoup(data, "html.parser")
                pre_tag = soup.select_one("div#jsonview pre")

                if pre_tag:
                    text_content = pre_tag.get_text(strip=True)
                    # print(text_content)  # 👉 {"hello": "world"}
                    tweet_json = json.loads(text_content)
                    tweet = parse_tweet_json(tweet_json, timestamp)
                    print(tweet)
                    update_xlsx(project_dir=project_dir,tweet=tweet)
                    if update_errors :
                        update_errors_xlsx(project_dir,tweet,file_number=file_number)
                    
                else:
                    tweet = parse_html(soup, link, True)
                    print(tweet)
                    update_xlsx(project_dir=project_dir,tweet=tweet)
                    if update_errors :
                        update_errors_xlsx(project_dir,tweet,file_number=file_number)
            else:
                tweet_json = json.loads(data)
                tweet = parse_tweet_json(tweet_json, timestamp)
                print(tweet)
                update_xlsx(project_dir=project_dir,tweet=tweet)
                if update_errors :
                    update_errors_xlsx(project_dir,tweet,file_number=file_number)
        else:
            soup = BeautifulSoup(data, "html.parser")
            tweet = parse_html(soup, file_name)
            update_xlsx(project_dir=project_dir,tweet=tweet)
            if update_errors :
                update_errors_xlsx(project_dir,tweet,file_number=file_number)
    except:
        if first_download == False:
            error = traceback.format_exc()
            print("---pre-saved download failed---")
            print(error)
            os.remove(index_path)
            download_with_wmd(url=url,timestamp=timestamp,project_dir=project_dir,user_name=user_name,content_type=content_type,output_dir=output_dir,update_errors=update_errors,file_number=file_number)  
        else:
            error = traceback.format_exc()
            print("---post-saved download also failed---")
            print(error)
            # Append only if not already in log
            if update_errors:
                try:
                    with open(os.path.join(project_dir, f"{project_dir}_errors_remaining_{file_number}.txt"), "r", encoding="utf-8") as f:
                        error_tweets = f.read().splitlines()
                except FileNotFoundError:
                    error_tweets = []

            # Use a set for faster lookups
                error_urls = set(error_tweets)
                if url not in error_urls:
                    error_urls.add(url)
                    error_tweets.append(url)
                    error_tweets.append(error)

                    with open(os.path.join(project_dir, f"{project_dir}_errors_remaining_{file_number}.txt"), "w", encoding="utf-8") as f:
                        f.write("\n".join(error_tweets))
                
            
            else:
                print(f"Error processing row: {e}")
                # traceback.print_exc()
                error_message = traceback.format_exc()
                print(error_message)

                # Append only if not already in log
                if url not in error_urls:
                    error_urls.add(url)
                    error_tweets.append(url)
                    error_tweets.append(error_message)

                    with open(os.path.join(project_dir, f"{project_dir}_error_tweets.txt"), "w", encoding="utf-8") as f:
                        f.write("\n".join(error_tweets))
                
    return True

def process_errors_tweets(tweet_array,project_dir,output_dir,file_number):
    print(tweet_array)
    print("Project_dir: ",project_dir)
    print("Output_dir: ",output_dir)
    
    
    original_url = tweet_array[0]
    timestamp = tweet_array[2]
    try:
        with open(os.path.join(project_dir, f"{project_dir}_errors_remaining_{file_number}.txt"), "r", encoding="utf-8") as f:
            error_tweets = f.read().splitlines()
    except FileNotFoundError:
        error_tweets = []

    # Use a set for faster lookups
    error_urls = set(error_tweets)
    
    try:
        if "application/json" in tweet_array[1]:
            download_with_wmd(
                url=original_url,
                timestamp=timestamp,
                project_dir=project_dir,
                content_type="application/json",
                output_dir=output_dir,
                user_name=project_dir,
                update_errors=True,
                file_number=file_number
            )
        else:
            download_with_wmd(
                url=original_url,
                timestamp=timestamp,
                project_dir=project_dir,
                output_dir=output_dir,
                user_name=project_dir,
                update_errors=True,
                file_number=file_number
            )
    except Exception as e:
        timestamp = tweet_array[3]
        try:
            if "application/json" in tweet_array[1]:
                download_with_wmd(
                    url=original_url,
                    timestamp=timestamp,
                    project_dir=project_dir,
                    content_type="application/json",
                    output_dir=output_dir,
                    user_name=project_dir,
                    update_errors=True,
                    file_number=file_number
                )
            else:
                download_with_wmd(
                    url=original_url,
                    timestamp=timestamp,
                    project_dir=project_dir,
                    output_dir=output_dir,
                    user_name=project_dir,
                    update_errors=True,
                    file_number=file_number
                )
        except:
            print(f"Error processing row: {e}")
            # traceback.print_exc()
            error_message = traceback.format_exc()
            print(error_message)

            # Append only if not already in log
            if original_url not in error_urls:
                error_urls.add(original_url)
                error_tweets.append(original_url)
                error_tweets.append(error_message)

                with open(os.path.join(project_dir, f"{project_dir}_errors_remaining_{file_number}.txt"), "w", encoding="utf-8") as f:
                    f.write("\n".join(error_tweets))

def process_json_file(json_path, project_dir, output_dir="archive", show_tqdm=True):
    # Load known error URLs (skip if file not found)
    try:
        with open(os.path.join(project_dir, f"{project_dir}_error_tweets.txt"), "r", encoding="utf-8") as f:
            error_tweets = f.read().splitlines()
    except FileNotFoundError:
        error_tweets = []

    # Use a set for faster lookups
    error_urls = set(error_tweets)

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Decide loop iterator
    iterator = tqdm(data, desc="Processing",file=sys.stderr) if show_tqdm else data

    # Skip header row
    for row in iterator:
        
        if not has_internet:
            print("Please check your internet connection")
            print("Sleeping for 10 seconds ...")
            time.sleep(10)
        
        original_url = row[0]
        if "status" not in original_url:
            continue 
        timestamp = row[2]

        if original_url in error_urls:
            print(f"Skipping {original_url}, already in error log.")
            continue

        try:
            if "application/json" in row[1]:
                download_with_wmd(
                    url=original_url,
                    timestamp=timestamp,
                    project_dir=project_dir,
                    content_type="application/json",
                    output_dir=output_dir,
                    user_name=project_dir,
                )
            else:
                download_with_wmd(
                    url=original_url,
                    timestamp=timestamp,
                    project_dir=project_dir,
                    output_dir=output_dir,
                    user_name=project_dir,
                )
        except Exception as e:
            timestamp = row[3]
            try:
                if "application/json" in row[1]:
                    download_with_wmd(
                        url=original_url,
                        timestamp=timestamp,
                        project_dir=project_dir,
                        content_type="application/json",
                        output_dir=output_dir,
                        user_name=project_dir,
                    )
                else:
                    download_with_wmd(
                        url=original_url,
                        timestamp=timestamp,
                        project_dir=project_dir,
                        output_dir=output_dir,
                        user_name=project_dir,
                    )
            except Exception as e:
                print(f"Error processing row: {e}")
                # traceback.print_exc()
                error_message = traceback.format_exc()
                print(error_message)

                # Append only if not already in log
                if original_url not in error_urls:
                    error_urls.add(original_url)
                    error_tweets.append(original_url)
                    error_tweets.append(error_message)

                    with open(os.path.join(project_dir, f"{project_dir}_error_tweets.txt"), "w", encoding="utf-8") as f:
                        f.write("\n".join(error_tweets))
        
    print("\n")


# Example usage
if __name__ == "__main__":
    # process_json_file("nekrovevo_captures.json", "archive")
    # files = os.listdir("archive")
    # print(f"The are {len(files)} files")
    # url = "https://twitter.com/nekrovevo/status/1283396073336143872"
    # timestamp = "20200715141220"
    # download_with_wmd(url=url,timestamp=timestamp,project_dir="Nekro",user_name="Nekro",output_dir="archive")
    # with open("index.html","r",encoding="utf-8") as file:
    #     soup = BeautifulSoup(file.read(),"html.parser")
    # tweet = parse_html(soup=soup,title="hello")
    # print(tweet)
    # output_dir = "archive"
    # tmp_dir = os.path.abspath(os.path.join(output_dir, "tmp_dl"))

    # if os.path.exists(tmp_dir):
    #     shutil.rmtree(tmp_dir)  # clean old runs
    # os.makedirs(tmp_dir, exist_ok=True)
    # https://web.archive.org/web/20200620164945/https://twitter.com/NekroVEVO/status/1274383951738503169
    download_with_wmd(url="https://twitter.com/NekroVEVO/status/1274383951738503169"
                      ,timestamp="20200620164945"
                      ,project_dir="NekroVevo"
                      ,user_name="NekroVevo"
                      ,output_dir="archive",
                      content_type="application/json")
